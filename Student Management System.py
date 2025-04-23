"""
Student Management System
Designed by Alick Li & Bryce Xing
"""

print("Student Management System")
print("Designed by Li Wantao & Xing Bowen & Yao Haowen")
print()
print("Loading...")

import tkinter as tk
from tkinter import Menu
from tkinter import messagebox
from tkinter import filedialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl
import numpy as np
from scipy import stats
from scipy.stats import skew
from scipy.stats import kurtosis

print("Done!")

# Create Root Window
root = tk.Tk()
root.title("Student Management System")
root.geometry("800x500")
root.resizable(False, False)

# Create a Label for the error message
labelError = tk.Label(root, text="", fg="red", anchor="w")
labelError.place(x=585, y=35, width=210, height=20)

# New Action, code for the "New" menubar command item, which clears the existing data and creates a new workspace.
def new_file():
    global students
    students = []
    global sorted_students
    sorted_students = []
    update_student_treeview()
    clear_entries()
    entryStatusBar.config(state='normal')
    entryStatusBar.delete(0, tk.END)
    entryStatusBar.insert(0, "")
    entryStatusBar.config(state='readonly')

# Open Action, code for the "Open" menubar command item, which opens an Excel file and reads the data into the program.
def open_file():
    entryStatusBar.config(state='normal')
    entryStatusBar.delete(0, tk.END)
    entryStatusBar.insert(0, "Loading...")
    entryStatusBar.config(state='readonly')
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        try:
            # Clear the existing treeview data
            students.clear()
            sorted_students.clear()
            clear_entries()
            update_student_treeview()
            # Open the Excel file and read the data
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            total_rows = worksheet.max_row-1
            for i in range(total_rows):
                student = {
                    "name": worksheet.cell(i+2,2).value,
                    "student_id": int(worksheet.cell(i+2,3).value),
                    "score": int(worksheet.cell(i+2,4).value),
                    "grade": calculate_grade(int(worksheet.cell(i+2,4).value))
                }
                students.append(student)
            sort_students(students, selectedSort.get())
            update_student_treeview()
            entryStatusBar.config(state='normal')
            entryStatusBar.delete(0, tk.END)
            entryStatusBar.insert(0, "Done!")
            entryStatusBar.config(state='readonly')
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
    def delay():
        entryStatusBar.config(state='normal')
        entryStatusBar.delete(0, tk.END)
        entryStatusBar.insert(0, "")
        entryStatusBar.config(state='readonly')
    entryStatusBar.after(1000, lambda: delay())

# Export Action, code for the "Export" menubar command item, which exports the data to an Excel file.
def export_file():
    entryStatusBar.config(state='normal')
    entryStatusBar.delete(0, tk.END)
    entryStatusBar.insert(0, "Loading...")
    entryStatusBar.config(state='readonly')
    from operator import indexOf
    from tkinter import filedialog
    try:
        file = openpyxl.Workbook()
        sheet = file.create_sheet("Sheet1", 0)
        title = ["No.", "Name", "Student ID", "Score", "Grade"]
        sheet["A1"] = title[0]
        sheet["B1"] = title[1]
        sheet["C1"] = title[2]
        sheet["D1"] = title[3]
        sheet["E1"] = title[4]
        for student in sorted_students:
            sheet.cell(row=indexOf(sorted_students, student)+2,column=1,value=indexOf(sorted_students, student)+1)
            sheet.cell(row=indexOf(sorted_students, student)+2,column=2,value=student["name"])
            sheet.cell(row=indexOf(sorted_students, student)+2,column=3, value=student["student_id"])
            sheet.cell(row=indexOf(sorted_students, student)+2,column=4, value=student["score"])
            sheet.cell(row=indexOf(sorted_students, student)+2,column=5, value=student["grade"])
        file.save(filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("XLSX files", "*.xlsx")]))
        file.close()
        entryStatusBar.config(state='normal')
        entryStatusBar.delete(0, tk.END)
        entryStatusBar.insert(0, "Done!")
        entryStatusBar.config(state='readonly')
    except Exception as e:
        pass
    def delay():
        entryStatusBar.config(state='normal')
        entryStatusBar.delete(0, tk.END)
        entryStatusBar.insert(0, "")
        entryStatusBar.config(state='readonly')
    entryStatusBar.after(1000, lambda: delay())

# Exit Action, which closes the program.
def exit():
    root.quit()

# 2 students lists, one for the original ones and one for the sorted ones.
students = []
sorted_students = []

# Add a new student to the list
def add_student():
    student = {
        "name": entryName.get(),
        "student_id": int(entryID.get()),
        "score": int(entryScore.get()),
        "grade": calculate_grade(int(entryScore.get()))
        }
    students.append(student)
    clear_entries()

# Confirm Action which is triggered by the "Confirm" button
def confirm_action():
    try:
        add_student()
        sort_students(students, selectedSort.get())
        update_student_treeview()
        labelError.config(text="")
    except Exception as e:
        labelError.config(text="Please enter valid information.")

# Delete selected student from the list
def delete_student():
    selected_item = treeviewStudents.selection()[0]
    student_id = int(treeviewStudents.item(selected_item, "values")[2])
    for student in students:
        if int(student['student_id']) == student_id:
            students.remove(student)
            break
    for student in sorted_students:
        if int(student['student_id']) == student_id:
            sorted_students.remove(student)
            break
    clear_entries()

# Delete Action which is triggered by the "Delete" button
def delete_action():
    try:
        delete_student()
        sort_students(students, selectedSort.get())
        update_student_treeview()
        labelError.config(text="")
    except Exception as e:
        labelError.config(text="Please select a student to delete.")

def draw_on_canvas():
    canvas.delete("all")
    try:
        bar_width = 355 / len(sorted_students) - 5
        bar_gap = 5
        for i, student in enumerate(sorted_students):
            score = student["score"]
            bar_height = score * 2.5
            x0 = i * (bar_width + bar_gap)
            y0 = 300 - bar_height
            x1 = x0 + bar_width
            y1 = 300
            canvas.create_rectangle(x0, y0, x1, y1, fill="blue")
            canvas.create_text(x0 + bar_width / 2, y0 - 10, text=str(score), anchor=tk.S)
    except Exception as e:
        pass

# Refresh the student treeview
def update_student_treeview():
    for item in treeviewStudents.get_children():
        treeviewStudents.delete(item)
    for student in sorted_students:
        treeviewStudents.insert("", "end", values=(student["number"], student["name"], student["student_id"], student["score"], student["grade"]))
    draw_on_canvas()
    statistics()
'''
def draw_on_canvas(canvas):

    input_list = [student["score"] for student in sorted_students]
    data = []
    for i in range(len(input_list)):
        data.append(input_list[i])
    input_array = [count_students()]
    fig, ax = plt.subplots()
    print(data)
    data = np.array(data)
    data = [88,75,85,92,68,91,67,93,86,54,85,95,84,96,67,97,77,78,81,99]
    plt.rcParams['font.sans-serif'] = ['SF Pro Display']
    plt.hist(data, bins=[0, 20, 40, 60, 80, 100], color='#FF7FB1')
    plt.title('Score Distribution Histogram')  # Title for the histogram
    # Create a canvas to embed the Matplotlib figure onto the existing one
    canvas_widget = FigureCanvasTkAgg(fig, master=canvas)
    canvas_widget.draw()
    canvas_widget.get_tk_widget().pack(fill=tk.BOTH, expand=True)
'''

def statistics():
    try:
        scores = []
        for student in sorted_students:
            scores.append(student["score"])
        mean = np.mean(scores)
        range = np.max(scores) - min(scores)
        median = np.median(scores)
        variance = np.var(scores)
        skewness = skew(np.array(scores))
        kurt = kurtosis(scores)
        entrymean.config(state='normal')
        entrymean.delete(0, tk.END)
        entrymean.insert(0, mean)
        entrymean.config(state='readonly')
        entrymedian.config(state='normal')
        entrymedian.delete(0, tk.END)
        entrymedian.insert(0, median)
        entrymedian.config(state='readonly')
        entryrange.config(state='normal')
        entryrange.delete(0, tk.END)
        entryrange.insert(0, range)
        entryrange.config(state='readonly')
        entryvariance.config(state='normal')
        entryvariance.delete(0, tk.END)
        entryvariance.insert(0, variance)
        entryvariance.config(state='readonly')
        entryskewness.config(state='normal')
        entryskewness.delete(0, tk.END)
        entryskewness.insert(0, skewness)
        entryskewness.config(state='readonly')
        entrykurtosis.config(state='normal')
        entrykurtosis.delete(0, tk.END)
        entrykurtosis.insert(0, kurt)
        entrykurtosis.config(state='readonly')
    except Exception as e:
        print(e)

def count_students():
    countA, countB, countC, countD, countF = 0,0,0,0,0
    for student in students:
        score = student["score"]
        if 90 <= score <= 100:
            countA += 1
        elif 80 <= score < 90:
            countB += 1
        elif 70 <= score < 80:
            countC += 1
        elif 60 <= score < 70:
            countD += 1
        elif 0 <= score < 60:
            countF += 1
    return countA,countB,countC,countD,countF


# Grade calculation
def calculate_grade(score):
    if 90 <= score <= 100:
        return 'A'
    elif 80 <= score < 90:
        return 'B'
    elif 70 <= score < 80:
        return 'C'
    elif 60 <= score < 70:
        return 'D'
    elif 0 <= score < 60:
        return 'F'
    else:
        raise Exception()
    

# Students sorting
def sort_students(students, selectedSort):
    if selectedSort == "Student ID   ↑":
        students.sort(key=lambda x: x["student_id"])
    elif selectedSort == "Score          ↓":
        students.sort(key=lambda x: x["score"], reverse=True)
    global sorted_students
    sorted_students = [{"number": i+1, **student} for i, student in enumerate(students)]
    return sorted_students

# Clear the entries
def clear_entries():
    entryNumber.config(state='normal')
    entryNumber.delete(0, 'end')
    entryNumber.config(state='readonly')
    entryName.delete(0, 'end')
    entryID.delete(0, 'end')
    entryScore.delete(0, 'end')
    entryGrade.config(state='normal')
    entryGrade.delete(0, 'end')
    entryGrade.config(state='readonly')

# Dropdown Select Event
def on_select(*args):
    sort_students(students, selectedSort.get())
    update_student_treeview()

# Treeview Select Event
def on_tree_select(event):
    selected_item = treeviewStudents.selection()[0]
    number = int(treeviewStudents.item(selected_item, "values")[0])
    sorted_students = sort_students(students, selectedSort.get())
    selected_student = sorted_students[number - 1]
    entryNumber.config(state='normal')
    entryNumber.delete(0, tk.END)
    entryNumber.insert(0, selected_student["number"])
    entryNumber.config(state='readonly')
    entryName.delete(0, tk.END)
    entryName.insert(0, selected_student["name"])
    entryID.delete(0, tk.END)
    entryID.insert(0, selected_student["student_id"])
    entryScore.delete(0, tk.END)
    entryScore.insert(0, selected_student["score"])
    entryGrade.config(state='normal')
    entryGrade.delete(0, tk.END)
    entryGrade.insert(0, selected_student["grade"])
    entryGrade.config(state='readonly')

# Create Menu
menubar = Menu(root)
menuFile = Menu(menubar, tearoff=0)
menuFile.add_command(label="New", command=new_file)
menuFile.add_command(label="Open...", command=open_file)
menuFile.add_command(label="Export...", command=export_file)
menuFile.add_command(label="Exit", command=exit)
menubar.add_cascade(label="File", menu=menuFile)
root.config(menu=menubar)

# Create Labels
labelNumber = tk.Label(root, text="No.", anchor="w")
labelNumber.place(x=15, y=15, width=50, height=20)
labelName = tk.Label(root, text="Name", anchor="w")
labelName.place(x=65, y=15, width=100, height=20)
labelStudentID = tk.Label(root, text="Student ID", anchor="w")
labelStudentID.place(x=165, y=15, width=150, height=20)
labelScore = tk.Label(root, text="Score", anchor="w")
labelScore.place(x=315, y=15, width=50, height=20)
labelGrade = tk.Label(root, text="Grade", anchor="w")
labelGrade.place(x=365, y=15, width=50, height=20)

labelTitle = tk.Label(root, text="Score Distribution Bar Graph")
labelTitle.place(x=430, y=85, width=355, height=20)

labelAvg = tk.Label(root, text="Average Score: ", anchor="w")
labelAvg.place(x=430, y=345, width=150, height=20)
labelMode = tk.Label(root, text="Mode Score: ", anchor="w")
labelMode.place(x=430, y=365, width=150, height=20)
labelRange = tk.Label(root, text="Score Range: ", anchor="w")
labelRange.place(x=430, y=385, width=150, height=20)
labelVariance = tk.Label(root, text="Score Variance: ", anchor="w")
labelVariance.place(x=430, y=405, width=150, height=20)
labelSkewness = tk.Label(root, text="Score Skewness: ", anchor="w")
labelSkewness.place(x=430, y=425, width=150, height=20)
labelKurtosis = tk.Label(root, text="Score Kurtosis: ", anchor="w")
labelKurtosis.place(x=430, y=445, width=150, height=20)

# Create Entries
entryNumber = tk.Entry(root,state="readonly")
entryNumber.place(x=15, y=35, width=50, height=20)
entryName = tk.Entry(root)
entryName.place(x=65, y=35, width=100, height=20)
entryID = tk.Entry(root)
entryID.place(x=165, y=35, width=150, height=20)
entryScore = tk.Entry(root)
entryScore.place(x=315, y=35, width=50, height=20)
entryGrade = tk.Entry(root,state="readonly")
entryGrade.place(x=365, y=35, width=50, height=20)

entrykurtosis = tk.Entry(root)
entrykurtosis.place(x=580, y=445, width=205, height=20)
entrykurtosis.config(state='readonly')
entryskewness = tk.Entry(root)
entryskewness.place(x=580, y=425, width=205, height=20)
entryskewness.config(state='readonly')
entryvariance = tk.Entry(root)
entryvariance.place(x=580, y=405, width=205, height=20)
entryvariance.config(state='readonly')
entryrange = tk.Entry(root)
entryrange.place(x=580, y=385, width=205, height=20)
entryrange.config(state='readonly')
entrymedian = tk.Entry(root)
entrymedian.place(x=580, y=365, width=205, height=20)
entrymedian.config(state='readonly')
entrymean = tk.Entry(root)
entrymean.place(x=580, y=345, width=205, height=20)
entrymean.config(state='readonly')

# Create Buttons
buttonConfirm = tk.Button(root, text="Confirm",command=confirm_action)
buttonConfirm.place(x=430, y=35, width=75, height=20)
buttonDelete = tk.Button(root, text="Delete",command=delete_action)
buttonDelete.place(x=505, y=35, width=75, height=20)

# Add Dropdown Menu
sortOptions = ["Student ID   ↑", "Score          ↓"]
selectedSort = tk.StringVar(root)
selectedSort.set(sortOptions[0])
dropdownSort = tk.OptionMenu(root, selectedSort, *sortOptions)
dropdownSort.place(x=428, y=53, width=154, height=24)

# Create Canvas
canvas = tk.Canvas(root, bg='white')
canvas.place(x=430, y=105, width=355, height=225)

# Create Treeview
from tkinter import ttk
treeviewStudents = ttk.Treeview(root, columns=("Number", "Name", "Student ID", "Score", "Grade"), show='headings')
treeviewStudents.place(x=15, y=55, width=400, height=410)
treeviewStudents.column("Number", width=50, anchor="w")
treeviewStudents.column("Name", width=100, anchor="w")
treeviewStudents.column("Student ID", width=150, anchor="w")
treeviewStudents.column("Score", width=50, anchor="w")
treeviewStudents.column("Grade", width=50, anchor="w")

# Bind Treeview Select Event
treeviewStudents.bind("<ButtonRelease-1>", on_tree_select)

# Bind Dropdown Select Event
selectedSort.trace("w", on_select)

# Create Status Bar
entryStatusBar = tk.Entry(root,state="readonly")
entryStatusBar.place(x=0, y=480, width=800, height=20)

root.mainloop()
